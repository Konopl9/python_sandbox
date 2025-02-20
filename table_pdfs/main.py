import openpyxl, os

workbook = openpyxl.load_workbook('test_python.xlsx')
print(workbook.sheetnames)
sheet = workbook['Sheet1']
cell = sheet['A1']
print(cell.value)
print(sheet.cell(row=1, column=1).value)
