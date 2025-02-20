import openpyxl

wb = openpyxl.Workbook()
print(wb.sheetnames)
sheet = wb['Sheet']
sheet.cell(row=1, column=1).value = 'Hello, world!'
print(sheet['A1'].value)
wb.save('New Excel File.xlsx')